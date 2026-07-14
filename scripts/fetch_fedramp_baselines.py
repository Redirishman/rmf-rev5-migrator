#!/usr/bin/env python3
"""Fetch and normalize the official FedRAMP Rev 5 control baselines.

FedRAMP baselines are *not* the NIST SP 800-53 baselines: FedRAMP selects a
superset (e.g. FedRAMP Moderate requires controls NIST Moderate does not), so a
FedRAMP package measured against a NIST baseline under-reports its gaps.

Source: FedRAMP's own "FedRAMP Security Controls Baseline" workbook, which
carries the High / Moderate / Low / LI-SaaS control selections on separate
sheets. FedRAMP retired the ``GSA/fedramp-automation`` OSCAL repository and the
``automate.fedramp.gov`` host during the 2026 consolidation; the workbook in the
FedRAMP ``docs-legacy`` repository is the surviving machine-readable publication
of the Rev 5 baselines and is what FedRAMP itself now links as legacy guidance.
It is stamped with a legacy notice dated 2026-06-23 — re-point ``WORKBOOK_URL``
if FedRAMP republishes the baselines in OSCAL.

Why bundle instead of fetch at runtime: private/GovCloud deployments have no
internet egress, so the tool ships this data inside the repo/artifact. A
maintainer runs this script (with internet) to (re)generate the committed files.

Usage:
    pip install -e "backend[dev]"   # provides openpyxl
    python scripts/fetch_fedramp_baselines.py
"""

from __future__ import annotations

import io
import json
import re
import sys
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

WORKBOOK_URL = (
    "https://raw.githubusercontent.com/FedRAMP/docs-legacy/main/"
    "overrides/assets/LEGACY%20FedRAMP_Security_Controls_Baseline.xlsx"
)

# sheet name -> output baseline name. The three impact baselines share a layout
# ("ID" column); LI-SaaS is a tailoring of Low and has its own layout.
IMPACT_SHEETS = {
    "Low Baseline": "fedramp_low",
    "Moderate Baseline": "fedramp_moderate",
    "High Baseline": "fedramp_high",
}
LI_SAAS_SHEET = "LI-SaaS Baseline"

ROOT = Path(__file__).resolve().parent.parent
BASELINE_DIR = ROOT / "data" / "baselines"
CATALOG_DIR = ROOT / "data" / "catalogs"

# "AC-2 (1)" / "ac-2(1)" / "AC-2" -> canonical "AC-2(1)" / "AC-2".
_CONTROL_RE = re.compile(r"^\s*([A-Za-z]{2})-(\d+)\s*((?:\(\s*\d+\s*\)\s*)*)$")
_ENHANCEMENT_RE = re.compile(r"\(\s*(\d+)\s*\)")


def _canonical_id(raw: str) -> str | None:
    """Normalize a workbook control id to the catalog's canonical form.

    Returns None for cells that are not control ids (section headers, notes,
    blank rows), so callers can filter them out.
    """
    match = _CONTROL_RE.match(raw)
    if not match:
        return None
    family, number, enhancements = match.groups()
    # Drop leading zeros so "AC-02" and "AC-2" collapse to one id.
    display = f"{family.upper()}-{int(number)}"
    for enhancement in _ENHANCEMENT_RE.findall(enhancements):
        display += f"({int(enhancement)})"
    return display


def _download(url: str) -> bytes:
    print(f"downloading {url}", file=sys.stderr)
    req = urllib.request.Request(url, headers={"User-Agent": "rmf-rev5-migrator"})  # noqa: S310
    with urllib.request.urlopen(req, timeout=120) as resp:  # noqa: S310
        return resp.read()


def _header_row(sheet: Any, wanted: str) -> tuple[int, int]:
    """Locate (row_index, column_index) of the column headed ``wanted``.

    The workbook puts a title banner above the header row, and the banner's
    depth differs per sheet, so the header is found rather than assumed.
    """
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
        for col_idx, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().casefold() == wanted.casefold():
                return row_idx, col_idx
    raise SystemExit(f"sheet {sheet.title!r}: no {wanted!r} column in the first 10 rows")


def parse_impact_sheet(sheet: Any) -> list[str]:
    """Control ids selected by a High/Moderate/Low FedRAMP baseline sheet."""
    header_idx, id_col = _header_row(sheet, "ID")
    ids: list[str] = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx <= header_idx or id_col >= len(row):
            continue
        cell = row[id_col]
        if isinstance(cell, str) and (control_id := _canonical_id(cell)):
            ids.append(control_id)
    return sorted(set(ids))


def parse_li_saas_sheet(sheet: Any) -> tuple[list[str], dict[str, str]]:
    """Control ids and per-control tailoring actions for the LI-SaaS baseline.

    LI-SaaS tailors the Low baseline: every control it lists is in scope, but the
    "Tailoring Action" column says how it is satisfied (Attest, Document and
    Assess, NSO — not system owner, etc.), which a reviewer needs to see.
    """
    header_idx, id_col = _header_row(sheet, "Control ID")
    _, action_col = _header_row(sheet, "Tailoring Action")
    ids: list[str] = []
    actions: dict[str, str] = {}
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx <= header_idx or id_col >= len(row):
            continue
        cell = row[id_col]
        if not isinstance(cell, str) or not (control_id := _canonical_id(cell)):
            continue
        ids.append(control_id)
        action = row[action_col] if action_col < len(row) else None
        if isinstance(action, str) and action.strip():
            actions[control_id] = action.strip()
    return sorted(set(ids)), actions


def _rev5_catalog_ids() -> set[str]:
    raw = json.loads((CATALOG_DIR / "rev5_controls.json").read_text())
    return {c["id"] for c in raw["controls"]}


def _write(name: str, payload: dict[str, Any]) -> None:
    (BASELINE_DIR / f"{name}.json").write_text(json.dumps(payload, indent=2) + "\n")


def main() -> int:
    BASELINE_DIR.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(io.BytesIO(_download(WORKBOOK_URL)), read_only=True)

    catalog_ids = _rev5_catalog_ids()
    counts: dict[str, int] = {}
    unknown_total = 0

    for sheet_name, baseline_name in IMPACT_SHEETS.items():
        ids = parse_impact_sheet(workbook[sheet_name])
        unknown = sorted(set(ids) - catalog_ids)
        unknown_total += len(unknown)
        if unknown:
            print(f"WARNING {baseline_name}: not in Rev 5 catalog: {unknown}", file=sys.stderr)
        _write(
            baseline_name,
            {
                "baseline": baseline_name,
                "source": WORKBOOK_URL,
                "source_sheet": sheet_name,
                "control_ids": ids,
            },
        )
        counts[baseline_name] = len(ids)

    ids, actions = parse_li_saas_sheet(workbook[LI_SAAS_SHEET])
    unknown = sorted(set(ids) - catalog_ids)
    unknown_total += len(unknown)
    if unknown:
        print(f"WARNING fedramp_li_saas: not in Rev 5 catalog: {unknown}", file=sys.stderr)
    _write(
        "fedramp_li_saas",
        {
            "baseline": "fedramp_li_saas",
            "source": WORKBOOK_URL,
            "source_sheet": LI_SAAS_SHEET,
            "control_ids": ids,
            "tailoring_actions": actions,
        },
    )
    counts["fedramp_li_saas"] = len(ids)

    print(f"wrote FedRAMP baselines {counts}", file=sys.stderr)
    return 1 if unknown_total else 0


if __name__ == "__main__":
    raise SystemExit(main())
